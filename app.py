from flask import Flask, render_template,request,redirect,url_for,jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import extract,func,or_
from sqlalchemy.sql.expression import extract
from openpyxl import load_workbook
from datetime import date,datetime
from dateutil.parser import parse
import logging
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///employer.db"
app.config['SQLALCHEMY_BINDS']={'login':"sqlite:///login.db",
                                'delete_user':"sqlite:///delete.db"
                                }

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

class Employee(db.Model):
    Sno = db.Column(db.Integer, primary_key=True, autoincrement=True)
    Emp_id = db.Column(db.String(500))
    Name = db.Column(db.String(500))
    Designation = db.Column(db.String(500))
    Department = db.Column(db.String(500))
    Project = db.Column(db.String(500))
    Job_role = db.Column(db.String(500))
    Employment_status = db.Column(db.String(500))
    Joining_date = db.Column(db.String(500))
    Experience = db.Column(db.String(500))
    Location = db.Column(db.String(500))
    Last_promoted = db.Column(db.String(500))
    Comments = db.Column(db.String(500))
class Login(db.Model):
    __bind_key__="login"
    id=db.Column(db.Integer,primary_key=True)
    email=db.Column(db.String(500))
    password=db.Column(db.String(200))
class Delete_user(db.Model):
    __bind_key__="delete_user"
    id=db.Column(db.Integer,primary_key=True) 
    Name=db.Column(db.String(200))
    Date=db.Column(db.String(200))   
def extract_data_from_excel():
    wb = load_workbook("employee_data 1.xlsx")
    ws = wb.active
    column_mappings = {
        'Sno': 0,
        'Emp_id': 1,
        'Name': 2,
        'Designation': 3,
        'Department': 4,
        'Project': 5,
        'Job_role': 6,
        'Employment_status': 7,
        'Joining_date': 8,
        'Experience': 9,
        'Location': 10,
        'Last_promoted': 11,
        'Comments': 12
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        
        if not all(cell is None for cell in row):
            Sno = row[column_mappings['Sno']]
            Emp_id = row[column_mappings['Emp_id']]
            Name = row[column_mappings['Name']]
            Designation = row[column_mappings['Designation']]
            Department = row[column_mappings['Department']]
            Project = row[column_mappings['Project']]
            Job_role = row[column_mappings['Job_role']]
            Employment_status = row[column_mappings['Employment_status']]
            Joining_date = row[column_mappings['Joining_date']]
            Experience = row[column_mappings['Experience']]
            formatted_date = None
            if isinstance(Joining_date, datetime):
                join_date = Joining_date 
                formatted_date=join_date.strftime("%d-%m-%Y")
                
                day = join_date.day
                month = join_date.month
                year = join_date.year

                current_date = date.today()
                Experience = current_date.year - year

                if (current_date.month, current_date.day) < (month, day):
                    Experience -= 1

                if Experience < 1:
                    Experience = "Less than 1 year"
            else:
                join_date = None
                day = None
                month = None
                year = None
                Experience = None

                 
            Location = row[column_mappings['Location']]
            Last_promoted = row[column_mappings['Last_promoted']]
            Comments = row[column_mappings['Comments']]
            


            existing_data = Employee.query.filter_by(Name=Name).first()
            if not existing_data:
                employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                    Department=Department, Project=Project, Job_role=Job_role,
                                    Employment_status=Employment_status, Joining_date=formatted_date,
                                    Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                    Comments=Comments)
                db.session.add(employee)
    db.session.commit()

def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1] in ["xlsx","csv"]
@app.route("/",methods=["GET","POST"])
def signPage():
    correct_user=None
    error_message=None
    if request.method=="POST":
        email=request.form["email"]
        password=request.form["password"]
        correct_user=Login.query.filter_by(email=email).first()
        if correct_user:
            if correct_user.password==password:
                return redirect(url_for("dashBoard"))
            else:
                correct_user=None
                error_message="invalid login credentials"
        if correct_user == None:
                error_message="invalid login credentials"      
    return render_template("sign.html",error_message=error_message)   
@app.route("/dashboard")
def dashBoard():
    status=Employee.query.with_entities(Employee.Employment_status).distinct()
    # for stat in status:
    #     count = Employee.query.filter_by(Employment_status=stat.Employment_status).count()
    #     print(f"Count for {stat.Employment_status}: {count}")
    # joining_dates = Employee.query.with_entities(Employee.Joining_date).all()
    # print(f"Joining dates: {joining_dates}")
    
    del_employers=Delete_user.query.all()
    deleted_jan_employers=[]
    deleted_feb_employers=[]
    deleted_march_employers=[]
    deleted_april_employers=[]
    deleted_may_employers=[]
    deleted_june_employers=[]
    deleted_july_employers=[]
    deleted_aug_employers=[]
    deleted_sep_employers=[]
    deleted_oct_employers=[]
    deleted_nov_employers=[]
    deleted_dec_employers=[]
    for emp in del_employers:
        if emp.Date:
            split_date=emp.Date.split("-")
            modified_month=split_date[1]
            if modified_month=="01":
                deleted_jan_employers.append(emp.Name) 
            if modified_month=="02":
                deleted_feb_employers.append(emp.Name) 
            if modified_month=="03":
                deleted_march_employers.append(emp.Name)
            if modified_month=="04":
                deleted_april_employers.append(emp.Name)
            if modified_month=="05":
                deleted_may_employers.append(emp.Name)                   
            if modified_month=="06":
                deleted_june_employers.append(emp.Name)
            if modified_month=="07":
                deleted_july_employers.append(emp.Name) 
            if modified_month=="08":
                deleted_aug_employers.append(emp.Name)
            if modified_month=="09":
                deleted_sep_employers.append(emp.Name)  
            if modified_month=='10':
                deleted_oct_employers.append(emp.Name)  
            if modified_month=='11':
                deleted_nov_employers.append(emp.Name)         
            if modified_month=='12':
                deleted_dec_employers.append(emp.Name)           
                 
                
                
    employers = Employee.query.all()
    jan_employers=[]
    feb_employers=[]
    march_employers=[]
    april_employers=[]
    may_employers=[]
    june_employers=[]
    july_employers=[]
    aug_employers=[]
    sep_employers=[]
    oct_employers=[]
    nov_employers=[]
    dec_employers=[]
    for employee in employers:
        if employee.Joining_date:
            split_date = employee.Joining_date.split('-')
            modified_month =split_date[1]
            # print("split dates")
            # print(modified_month)            
            if modified_month =="01":
                jan_employers.append(employee.Name)  
            if modified_month =="02":
                feb_employers.append(employee.Name)    
            if modified_month =="03":
                march_employers.append(employee.Name)
            if modified_month =="04":
                april_employers.append(employee.Name)           
            if modified_month =="05":
                may_employers.append(employee.Name)
            if modified_month=='06':
               june_employers.append(employee.Name) 
            if modified_month=='07':
               july_employers.append(employee.Name)      
            if modified_month=='08':
               aug_employers.append(employee.Name) 
            if modified_month=='09':
               sep_employers.append(employee.Name)
            if modified_month=='10':
               oct_employers.append(employee.Name)
            if modified_month=='11':
               nov_employers.append(employee.Name)
            if modified_month=='12':
               dec_employers.append(employee.Name)               
    # print("june employers..")
    # print(june_employers)

    employment_status_counts={}
    for stat in status:
        count=Employee.query.filter_by(Employment_status=stat.Employment_status).count()
        
        employment_status_counts[stat.Employment_status]=count
    return render_template("dashboard.html",employment_status_counts=employment_status_counts,june_employers=june_employers,deleted_june_employers=deleted_june_employers,deleted_jan_employers=deleted_jan_employers,deleted_feb_employers=deleted_feb_employers,deleted_march_employers=deleted_march_employers,deleted_april_employers=deleted_april_employers,deleted_may_employers=deleted_may_employers,deleted_july_employers=deleted_july_employers,deleted_aug_employers=deleted_aug_employers,deleted_sep_employers=deleted_sep_employers,deleted_oct_employers=deleted_oct_employers,deleted_nov_employers=deleted_nov_employers,deleted_dec_employers=deleted_dec_employers,jan_employers=jan_employers,feb_employers=feb_employers,march_employers=march_employers,april_employers=april_employers,may_employers=may_employers,july_employers=july_employers,aug_employers=aug_employers,sep_employers=sep_employers,oct_employers=oct_employers,nov_employers=nov_employers,dec_employers=dec_employers) 
@app.route("/home")
def Home():
    data=Employee.query.all()
    return render_template("index.html",data=data)

@app.route("/add", methods=["GET", "POST"])
def Add():
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        date_parts = joining_date.split('-')
        if len(date_parts) == 3:
            formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            join_date=datetime.strptime(formatted_date, "%d-%m-%Y")
            current_date=date.today()   
            experience=current_date.year -int(join_date.year)
            if (current_date.month,current_date.day) < (join_date.month,join_date.day):
                experience-=1
            if experience <1:
                experience="Less than 1 year"       
        else:
            formatted_date = None 
            experience=None
        # experience = request.form.get("experience")
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        existing_data=Employee.query.filter_by(Name=name).first()
        if not existing_data:
            employee = Employee(
                Emp_id=emp_id,
                Name=name,
                Designation=designation,
                Department=department,
                Project=project,
                Job_role=job_role,
                Employment_status=employment_status,
                Joining_date=formatted_date,
                Experience=experience,
                Location=location,
                Last_promoted=last_promoted,
                Comments=comments
            )
            db.session.add(employee)
            db.session.commit()
        return redirect("/home")
    return render_template("add.html")

@app.route("/update/<int:sno>",methods=["GET","POST"])
def Update(sno):
    selected_date = request.args.get("date")
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        date_parts=joining_date.split("-")
        if len(date_parts)==3:
            formatted_date=f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            join_date=datetime.strptime(formatted_date,"%d-%m-%Y")
            current_day=date.today()
            experience=current_day.year-int(join_date.year)
            if (current_day.month,current_day.day) < (join_date.month,join_date.day):
                experience-=1
            if experience < 1:
                experience="Less than 1 year"
        else:
            formatted_date = None 
            experience=None            
        # experience = request.form.get("experience")
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        employee=Employee.query.filter_by(Sno=sno).first()
        employee.Emp_id=emp_id
        employee.Name=name
        employee.Designation=designation
        employee.Department=department
        employee.Project=project
        employee.Job_role=job_role
        employee.Employment_status=employment_status
        employee.Joining_date=formatted_date
        employee.Experience=experience
        employee.Location=location
        employee.Last_promoted=last_promoted
        employee.Comments=comments
        db.session.add(employee)
        db.session.commit()
        return redirect("/home")
    employee=Employee.query.filter_by(Sno=sno).first()
    return render_template("update.html",employee=employee,selected_date=selected_date)

@app.route("/delete/<int:sno>")
def Delete(sno):
    employee=Employee.query.filter_by(Sno=sno).first()
    delete=Delete_user(Name=employee.Name,Date=employee.Joining_date)
    db.session.add(delete)
    db.session.commit()
    db.session.delete(employee)
    db.session.commit()
    return redirect("/home")
with app.app_context():
        db.create_all()
        data=extract_data_from_excel()

@app.route("/bulk",methods=["GET","POST"])
def bulk():
    if request.method=="POST":
        file = request.files['file']
        if file and allowed_file(file.filename):
            if file.filename.endswith(".xlsx"):
                wb=load_workbook(file)
                ws=wb.active
                column_mappings = {
                    'Sno': 0,
                    'Emp_id': 1,
                    'Name': 2,
                    'Designation': 3,
                    'Department': 4,
                    'Project': 5,
                    'Job_role': 6,
                    'Employment_status': 7,
                    'Joining_date': 8,
                    'Experience': 9,
                    'Location': 10,
                    'Last_promoted': 11,
                    'Comments': 12
                    }
                for row in ws.iter_rows (min_row=2,values_only=True):
                    if not all(cell is None for cell in row):
                        
                        Emp_id = row[column_mappings['Emp_id']]
                        Name = row[column_mappings['Name']]
                        Designation = row[column_mappings['Designation']]
                        Department = row[column_mappings['Department']]
                        Project = row[column_mappings['Project']]
                        Job_role = row[column_mappings['Job_role']]
                        Employment_status = row[column_mappings['Employment_status']]
                        Joining_date = row[column_mappings['Joining_date']]
                        Experience = row[column_mappings['Experience']]
                        formatted_date = None
                        if isinstance(Joining_date,datetime):
                            join_date=Joining_date
                            formatted_date=join_date.strftime("%d-%m-%Y")
                            month=join_date.month
                            day=join_date.day
                            year=join_date.year
                            current_date=date.today()
                            Experience=current_date.year-year
                            if (current_date.month,current_date.day) < (month,day):
                                Experience-=1
                            if Experience < 1:
                                Experience="Less than 1 year"
                        else:
                            join_date = None
                            day = None
                            month = None
                            year = None
                            Experience = None

                 
                        Location = row[column_mappings['Location']]
                        Last_promoted = row[column_mappings['Last_promoted']]
                        Comments = row[column_mappings['Comments']]
                        existing_data=Employee.query.filter_by(Name=Name).first()
                        if not existing_data:
                            employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                    Department=Department, Project=Project, Job_role=Job_role,
                                    Employment_status=Employment_status, Joining_date=formatted_date,
                                    Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                    Comments=Comments)
                            db.session.add(employee)
                db.session.commit()
                return redirect("/home")            

    
    return render_template("bulk.html")
    
@app.route("/view/<int:sno>")
def view(sno):
    data=Employee.query.filter_by(Sno=sno).first()
    return render_template("view.html",data=data)    
@app.route("/register",methods=["GET","POST"])
def register():
    if request.method=="POST":
        email=request.form["email"]
        password=request.form["password"]
        user=Login(email=email,password=password)
        db.session.add(user)
        db.session.commit()
        return redirect("/")

    return render_template("register.html")
@app.route("/get_employees_list/<employment_status>")
def get_employees_list(employment_status):
    employees=Employee.query.filter_by(Employment_status=employment_status).all()
    employee_names=[employee.Name for employee in employees]
    return jsonify({'employeeList': employee_names})

if __name__ == "__main__":
    app.run(debug=True)



